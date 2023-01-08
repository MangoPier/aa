'''
t1 = (10,)
print(type(t1))
t2= ('zfj','lucas',[1,2,3])
print(type(t2))
print(t2.index('zfj'))
t2[2][1]=4
print(t2)
'''


# 外部数据源文件处理
'''
# 1、Excel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
wb = Workbook()
dest_filename = 'my_sheet.xlsx'

ws1 = wb.active # 调用一个正在用的Excel表
ws1.title = "the first sheet"

for i  in range(1,40):
    ws1.append(range(10)) # 39*9的数据

ws2 = wb.create_sheet(title="the second sheet")

ws2['G20'] = 201314 # 对某些单元格写入值

ws3 = wb.create_sheet(title="the third sheet")
for row in range(10,20):
    for col in range(27,54):
        _ = ws3.cell(column=col,row=row,value="{0}".format(get_column_letter(col))) # 获取到列的字母
print(ws3['AA10'].value)

ws4 = wb.create_sheet(title="the forth sheet")
for i in range(1,31):
    ws4.cell(column=1,row=i).value=f"test{i}"
wb.save(filename=dest_filename)

from openpyxl import load_workbook
wb = load_workbook(filename = 'my_sheet.xlsx')
sheet_ranges = wb['the first sheet']
print(sheet_ranges['D18'].value)
for i in range(1,31):
    print(sheet_ranges.cell(column=1,row=i).value)

'''

# 2、yaml——可以进行格式转换，关键是横杠，有横杆就是列表
"""

import yaml
# print(yaml.load("""
# - 1
# - 2
# - 3""",Loader=yaml.FullLoader))
# 转换为列表：[1, 2, 3]

# print(yaml.load("""
# a: 1
# """,Loader=yaml.FullLoader))
# 转换为字典：{'a': 1}

# print(yaml.load(open("demo.yaml"),Loader=yaml.FullLoader))
# {'a': [1, 2]}，字典嵌套列表
"""

with open("demo3.yml","w") as f:
    print(yaml.dump({'a': [1, 2]},stream=f))
# 把内容写入到demo3，需要写入的权限

a = "aaaa"
b = "bbbb"
c = 1+2j # 复数
print(type(c)) # <class 'complex'>
print(a+b)
print(b.format(a)) # bbbb

x = 0
if x > 1:
    y = 3*x +5
else:
    if x >= -1:
        y = x +2
    else:
        y = 5*x +3
print(y)
# 扁平化结构优于嵌套结构
result = 0
for i in range(101):
    if i%2 ==0:
        print(i)
        result = result + i
print(result)
a = 1
while a == 1:
    print("a==1")
    a = a +1
else:
    print("a != 1")
    print(a)

for i in range(1,10):
    if i == 5:
        break
    print(i)
"""


# 猜数字
"""

import random
person_number = 0
computer_number = random.randint(1,100)
while True:
    person_number = int(input("请输入整数："))
    if person_number>computer_number:
        print("小一点")
    elif person_number < computer_number:
        print("大一点")
    else:
        print("猜对啦")
        break
"""

"""
def method(a): # a是必须参数
    print(a)
    return a + 2

method(1)
# 返回1
print("----------")
print(method(1))
# 返回1、3

def method2(a,b=[]):
    b.append(a)
    return b

print(method2(2)) # [2]
print(method2(3)) # [2, 3]
# 默认值只执行一次

# 关键字参数，一定要在默认参数后面
def method3(a ,b):
    print(a)
    print(b)
    return b
# print(method3(a=2,3))  # 报错。关键字参数，一定要在默认参数后面
print(method3(1,2))
print(method3(a = 3,b=4))
print(method3(5,b=6))

# 特殊参数：*仅限关键字参数，在【仅限关键字】形参前放置一个*
def method4(**a): # **a代表传入参数为字典；*a是元组传参
    print(a.keys())

method4(a =1,b=2,c=3) #dict_keys(['a', 'b', 'c'])
# method4(1,2,3) #会报错

def method5(*a):
    print(a[0])
method5(1,2,3,4,5) #1

# 解包参数列表
# *用来解包元组；**用来解包字典
print(list(range(3,6)))

list_a = (3,6)
# list(range(list_a)) # TypeError: 'tuple' object cannot be interpreted as an integer
list(range(*list_a))  # 用*给元组解包，[3, 4, 5]

def method6(a,b,c):
    print(a)
    print(b)
    print(c)
dict1 = {"a":1,"b":2,"c":3}
# method6(dict1) #报错 TypeError: method6() missing 2 required positional arguments: 'b' and 'c'
method6(**dict1) # 用**给字典解包，1、2、3

# lambda关键字可以用于创建一个小的匿名函数，关键字为lambda，lambda的主体是一个表达式，而不是一个代码块，仅仅能在lambda表达式中封装有限的逻辑进去
y = lambda x,y,z:x+y+z
print(y(2,3,4)) # 9
"""

"""
print("-------------------------这里是复习字面量插值------------------------")
list = [1,2,3]
dict = {"a":1,"b":2}
print("we are {} and {}".format("brave","rich"))
print("these numbers are {},{},{}".format(*list))
print("a is {a} and b is {b}".format(**dict))
name = 'zfj'
print(f"my name is {name.upper()}")
"""
"""
print("-------------------------这里是复习文件打开------------------------")
with open("data.yaml",'r') as f:
    print(f.read())
    print("------------")
    print(f.readline())
    print(f.readline())
    print("-------------")
    print(f.readlines())
    print("-----------")
    f.close()
    print(f.readable())
"""
"""
print("______________________这里是复习json______________________________")
import json

# json.dumps把数据类型转换为字符串，支持数字、列表、字典、元组。但是不支持集合
num1 = 1
num1 = [1,2,3]
num1 = {"a":1,"b":2}
num1 = (1,2,3)
j1 = json.dumps(num1) # 把数据类型转换为字符串
print(j1)
print(type(j1))
# <class 'str'>

# json.loads可以把json字符串(字符串里面是字典或者列表）转换为字典或列表
j2 = json.loads('{"a":1}')
print(j2)
print(type(j2))
# <class 'dict'>

j2 = json.loads('[1,2]')
print(j2)
print(type(j2))
# <class 'list'>

j3 = json.loads('[{"a":1},{"b":2,"c":3}]')
print(j3)
print(type(j3))
# <class 'list'>


# 以下情况是不允许用于json.loads
# j3 = json.loads('(1,2)')
# j3 = json.loads("1,2")
# print(j3)
# print(type(j3))


# json.load方法，支持从json文件中读取数据
with open("jsontest.json",'r') as f:
    print(json.load(f))
    # [{'name': 'tom', 'gender': 'male', 'other': None}, {'name': 'lucas', 'gender': 'female', 'other': None}]

print("------------------------------------------------")

# json.dump可以把数据以json的数据类型写入文件中。这种数据可以是字典、列表、元组、数字（整型、浮点型）、字符串

dict1 = [1,2]
dict1 = (2,2)
dict1 = 1.4
dict1 = {"name":"zfj","age":26}
dict2 = "hhh"
dict3 = "lala"
# dict1 = input("请输入一串文字用于写入jsfile.json文件中：")
with open("jsfile.json",'w') as f:
    print("----------这里用dump方法给文件写入数据-----------")
    json.dump(dict1,f,indent=6)
    print("----------dump方法写入文件已完成-----------")


with open("jsfile.json",'r') as f:
    print(json.load(f))
    # {'name': 'zfj', 'age': 26}
"""

print("-------------------------这里是复习异常与错误-----------------------------")
try:
    assert 1 == 2
    print("try-------")
except:
    print("上述代码运行有异常，这里是一个except")
else:
    print("如果没有异常就会执行else")
finally:
    print("这是一个finally，不论有没有异常都会执行")






